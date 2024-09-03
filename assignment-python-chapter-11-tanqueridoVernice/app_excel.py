""" Good references for more info: 
https://openpyxl.readthedocs.io/en/stable/index.html
http://zetcode.com/python/openpyxl/
"""

import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)
# Open a sheet to work with
sheet = wb["Sheet1"]
wb.create_sheet("Sheet2")

# Reference cells on a sheet
cell = sheet["a1"]
print(cell.value)
cell.value = "Transaction ID"
print(cell.row)
print(cell.column)
print(cell.coordinate)
cell = sheet.cell(row=1, column=1)

print("-"*30)
# Loop through various cells
for c in range(1, sheet.max_column + 1):
    for r in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=r, column=c)
        print(cell.value)

print("-"*30)

# Reference an entire column of cells
# column is a tuple of cells
column = sheet["a"]
print(column)
for cell in column:
    print(cell.value)

print("-"*30)
# Reference an entire column of cells
# columns ia a tuple of cells
columns = sheet["a:c"]
print(columns)
for column in columns:
    for cell in column:
        print(cell.value)

range = sheet["a1:c3"]
first_row = sheet[1]
three_rows = sheet[1:3]

# Append one row of data
# list or tuple will work
data = [1004, 4, 9.99]
sheet.append(data)

# Append several rows of data
# lists or tuples wil work
new_rows_data = [
    [1004, 4, 9.99],
    [1005, 5, 7.52],
    [1006, 6, 1.99]
]

for row in new_rows_data:
    sheet.append(row)

# Save the workbook as new file
wb.save("transactions2.xlsx")
